#!/usr/bin/env python3
"""
KIS OpenAPI Excel Parsing and Documentation Generation Script

Parses HTS_OPENAPI.xlsx "API 목록" sheet to generate TR_ID mapping table
and individual API documentation files.

Usage:
    # Parse Excel and generate TR_ID mapping
    python scripts/parse_kis_excel.py

    # Generate API documentation files
    python scripts/parse_kis_excel.py --generate-docs

    # Parse Excel and generate documentation
    python scripts/parse_kis_excel.py --all

Output:
    docs/kis-openapi/_data/tr_id_mapping.json - TR_ID mapping table
    docs/kis-openapi/api/*.md - Individual API documentation files
"""

import argparse
import json
import logging
import re
from pathlib import Path
from typing import Dict, Any, List, Optional

import openpyxl
from jinja2 import Environment, FileSystemLoader, Template

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# File paths
PROJECT_ROOT = Path(__file__).parent.parent
EXCEL_FILE = PROJECT_ROOT / "HTS_OPENAPI.xlsx"
DOCS_RAW_DIR = PROJECT_ROOT / "docs_raw"
CATEGORIES_FILE = DOCS_RAW_DIR / "categories.json"
TEMPLATE_FILE = DOCS_RAW_DIR / "template.md"
OUTPUT_DIR = PROJECT_ROOT / "docs" / "kis-openapi" / "_data"
API_DOCS_DIR = PROJECT_ROOT / "docs" / "kis-openapi" / "api"


# Category slug mapping (Korean to English slug)
CATEGORY_SLUGS = {
    "OAuth인증": "oauth-authentication",
    "[국내주식] 기본시세": "domestic-stock-basic",
    "[국내주식] 실시간시세": "domestic-stock-realtime",
    "[국내주식] 주문/계좌": "domestic-stock-orders",
    "[국내주식] 시세분석": "domestic-stock-analysis",
    "[국내주식] 순위분석": "domestic-stock-ranking",
    "[국내주식] 종목정보": "domestic-stock-info",
    "[국내주식] ELW 시세": "domestic-stock-elw",
    "[국내주식] 업종/기타": "domestic-stock-sector",
    "[국내선물옵션] 기본시세": "domestic-futures-basic",
    "[국내선물옵션] 실시간시세": "domestic-futures-realtime",
    "[국내선물옵션] 주문/계좌": "domestic-futures-orders",
    "[장내채권] 기본시세": "bond-basic",
    "[장내채권] 실시간시세": "bond-realtime",
    "[장내채권] 주문/계좌": "bond-orders",
    "[해외주식] 기본시세": "overseas-stock-basic",
    "[해외주식] 실시간시세": "overseas-stock-realtime",
    "[해외주식] 주문/계좌": "overseas-stock-orders",
    "[해외주식] 시세분석": "overseas-stock-analysis",
    "[해외선물옵션] 기본시세": "overseas-futures-basic",
    "[해외선물옵션]실시간시세": "overseas-futures-realtime",
    "[해외선물옵션] 주문/계좌": "overseas-futures-orders",
}


def sanitize_filename(api_id: str) -> str:
    """
    Sanitize API ID for use as filename.

    Args:
        api_id: Raw API ID from Excel

    Returns:
        str: Sanitized filename-safe string
    """
    # Replace special characters with hyphens
    sanitized = re.sub(r'[^\w\-]', '-', api_id)
    # Remove consecutive hyphens
    sanitized = re.sub(r'-+', '-', sanitized)
    # Remove leading/trailing hyphens
    sanitized = sanitized.strip('-')
    return sanitized


def parse_excel() -> Dict[str, Any]:
    """
    Parse HTS_OPENAPI.xlsx "API 목록" sheet.

    Returns:
        Dict[str, Any]: TR_ID mapping dictionary

    Structure:
        {
            "api_id": {
                "name": "API Name",
                "real_tr_id": "Real Trading TR_ID",
                "paper_tr_id": "Paper Trading TR_ID",
                "category": "Menu Location",
                "http_method": "GET/POST",
                "url": "/uapi/...",
                "communication_type": "REST/WEBSOCKET"
            }
        }
    """
    logger.info(f"Parsing Excel file: {EXCEL_FILE}")

    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")

    # Load workbook
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # Get "API 목록" sheet
    if "API 목록" not in wb.sheetnames:
        raise ValueError("'API 목록' sheet not found in Excel file")

    ws = wb["API 목록"]

    # Initialize mapping dictionary
    tr_id_mapping: Dict[str, Any] = {}
    api_count = 0
    real_tr_id_count = 0
    paper_tr_id_count = 0

    # Iterate through rows (skip header row)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Extract columns
        # Row structure:
        # 0: 순번, 1: API 통신방식, 2: 메뉴 위치, 3: API 명,
        # 4: API ID, 5: 실전 TR_ID, 6: 모의 TR_ID,
        # 7: HTTP Method, 8: URL 명, 9: 실전 Domain, 10: 모의 Domain
        seq_no = row[0]
        comm_type = row[1]
        menu_location = row[2]
        api_name = row[3]
        api_id = row[4]
        real_tr_id = row[5]
        paper_tr_id = row[6]
        http_method = row[7]
        url = row[8]

        # Skip if api_id is missing
        if not api_id:
            logger.warning(f"Row {row_idx}: Missing API ID, skipping")
            continue

        # Create API entry
        api_entry = {
            "name": api_name or "",
            "real_tr_id": real_tr_id or "",
            "paper_tr_id": paper_tr_id or "",
            "category": menu_location or "",
            "http_method": http_method or "",
            "url": url or "",
            "communication_type": comm_type or "",
        }

        # Add to mapping
        tr_id_mapping[api_id] = api_entry
        api_count += 1

        # Track TR_ID counts
        if real_tr_id:
            real_tr_id_count += 1
        if paper_tr_id:
            paper_tr_id_count += 1

    logger.info(f"Parsed {api_count} APIs")
    logger.info(f"  - Real TR_IDs: {real_tr_id_count}")
    logger.info(f"  - Paper TR_IDs: {paper_tr_id_count}")

    return tr_id_mapping


def validate_json_schema(data: Dict[str, Any]) -> bool:
    """
    Validate JSON schema.

    Args:
        data: TR_ID mapping dictionary

    Returns:
        bool: True if valid
    """
    logger.info("Validating JSON schema...")

    # Check if data is a dictionary
    if not isinstance(data, dict):
        raise ValueError("Data must be a dictionary")

    # Check if data is not empty
    if not data:
        raise ValueError("Data is empty")

    # Validate each API entry
    for api_id, entry in data.items():
        if not isinstance(entry, dict):
            raise ValueError(f"API {api_id}: Entry must be a dictionary")

        # Required fields
        required_fields = ["name", "real_tr_id", "paper_tr_id", "category", "http_method", "url", "communication_type"]
        for field in required_fields:
            if field not in entry:
                raise ValueError(f"API {api_id}: Missing required field '{field}'")

    logger.info("JSON schema validation passed")
    return True


def save_json(data: Dict[str, Any]) -> None:
    """
    Save JSON data to file.

    Args:
        data: TR_ID mapping dictionary
    """
    logger.info(f"Saving JSON to: {OUTPUT_DIR}")

    # Create output directory if it doesn't exist
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    output_file = OUTPUT_DIR / "tr_id_mapping.json"

    # Save with pretty formatting
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    logger.info(f"JSON saved successfully ({len(data)} APIs)")


def load_template() -> Template:
    """
    Load Jinja2 template for documentation generation.

    Returns:
        Template: Jinja2 template object

    Raises:
        FileNotFoundError: If template file not found
    """
    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"Template file not found: {TEMPLATE_FILE}")

    logger.info(f"Loading template: {TEMPLATE_FILE}")

    # Create Jinja2 environment
    env = Environment(
        loader=FileSystemLoader(str(DOCS_RAW_DIR)),
        autoescape=False
    )

    # Load template
    template = env.get_template("template.md")

    logger.info("Template loaded successfully")
    return template


def categorize_apis(tr_id_mapping: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Categorize APIs by their category slug.

    Args:
        tr_id_mapping: TR_ID mapping dictionary

    Returns:
        Dict[str, List[Dict]]: Categorized APIs
    """
    categorized = {}

    for api_id, api_data in tr_id_mapping.items():
        category = api_data["category"]
        category_slug = CATEGORY_SLUGS.get(category, "other")

        if category_slug not in categorized:
            categorized[category_slug] = []

        categorized[category_slug].append({
            "id": sanitize_filename(api_id),
            "api_id": api_id,
            **api_data
        })

    # Log category distribution
    logger.info(f"APIs categorized into {len(categorized)} groups:")
    for category_slug, apis in sorted(categorized.items()):
        logger.info(f"  - {category_slug}: {len(apis)} APIs")

    return categorized


def generate_api_docs(
    tr_id_mapping: Dict[str, Any],
    phase: Optional[int] = None,
    priority_categories: Optional[List[str]] = None
) -> None:
    """
    Generate individual API documentation files.

    Args:
        tr_id_mapping: TR_ID mapping dictionary
        phase: Generation phase (1 = priority APIs, 2 = remaining APIs, None = all)
        priority_categories: List of category slugs for phase 1 (priority)
    """
    logger.info("Generating API documentation files...")

    # Load template
    template = load_template()

    # Categorize APIs
    categorized_apis = categorize_apis(tr_id_mapping)

    # Filter based on phase
    if phase == 1 and priority_categories:
        logger.info(f"Phase {phase}: Generating priority APIs only")
        filtered_apis = {
            cat: apis for cat, apis in categorized_apis.items()
            if cat in priority_categories
        }
    elif phase == 2 and priority_categories:
        logger.info(f"Phase {phase}: Generating remaining APIs")
        filtered_apis = {
            cat: apis for cat, apis in categorized_apis.items()
            if cat not in priority_categories
        }
    else:
        logger.info("Generating all APIs (all phases)")
        filtered_apis = categorized_apis

    # Generate documentation files
    total_generated = 0

    for category_slug, apis in filtered_apis.items():
        # Create category directory
        category_dir = API_DOCS_DIR / category_slug
        category_dir.mkdir(parents=True, exist_ok=True)

        # Create category index file
        index_file = category_dir / "index.md"
        generate_category_index(index_file, category_slug, apis)

        # Generate individual API documentation
        for api in apis:
            output_file = category_dir / f"{api['id']}.md"

            # Prepare template context
            context = {
                "api_id": api["api_id"],
                "api_name": api["name"],
                "category": api["category"],
                "http_method": api["http_method"],
                "url": api["url"],
                "communication_type": api["communication_type"],
                "real_tr_id": api["real_tr_id"],
                "paper_tr_id": api["paper_tr_id"],
                "has_params": api["http_method"] in ["GET", "POST"],
                "related_apis": []  # Could be enhanced with related API logic
            }

            # Render template
            content = template.render(**context)

            # Write to file
            with open(output_file, "w", encoding="utf-8") as f:
                f.write(content)

            total_generated += 1

    logger.info(f"Generated {total_generated} API documentation files")
    logger.info(f"Documentation directory: {API_DOCS_DIR}")


def generate_category_index(index_file: Path, category_slug: str, apis: List[Dict[str, Any]]) -> None:
    """
    Generate category index file.

    Args:
        index_file: Path to index file
        category_slug: Category slug
        apis: List of APIs in this category
    """
    # Read category name from first API
    category_name = apis[0]["category"] if apis else category_slug

    # Generate index content
    content = f"# {category_name}\n\n"
    content += f"Total APIs: {len(apis)}\n\n"
    content += "## API List\n\n"

    for api in sorted(apis, key=lambda x: x["name"]):
        content += f"- [{api['name']}]({api['id']}.md) - `{api['api_id']}`\n"

    content += "\n---\n\n"
    content += "[Back to API Documentation](../index.md)\n"

    # Write index file
    with open(index_file, "w", encoding="utf-8") as f:
        f.write(content)

    logger.info(f"Generated category index: {index_file}")


def generate_main_index(tr_id_mapping: Dict[str, Any]) -> None:
    """
    Generate main API documentation index.

    Args:
        tr_id_mapping: TR_ID mapping dictionary
    """
    logger.info("Generating main API index...")

    categorized_apis = categorize_apis(tr_id_mapping)

    # Load categories info
    categories_info = {}
    if CATEGORIES_FILE.exists():
        with open(CATEGORIES_FILE, "r", encoding="utf-8") as f:
            categories_data = json.load(f)
            for cat in categories_data.get("categories", []):
                categories_info[cat["id"]] = cat

    # Generate index content
    content = "# KIS OpenAPI Documentation\n\n"
    content += f"Total APIs: {len(tr_id_mapping)}\n\n"
    content += "## Categories\n\n"

    for category_slug in sorted(categorized_apis.keys()):
        cat_info = categories_info.get(category_slug, {})
        cat_name = cat_info.get("name", category_slug)
        cat_desc = cat_info.get("description", "")
        api_count = len(categorized_apis[category_slug])

        content += f"### [{cat_name}](api/{category_slug}/index.md)\n\n"
        if cat_desc:
            content += f"{cat_desc}\n\n"
        content += f"APIs: {api_count}\n\n"

    content += "---\n\n"
    content += "## API Documentation\n\n"

    for category_slug in sorted(categorized_apis.keys()):
        cat_info = categories_info.get(category_slug, {})
        cat_name = cat_info.get("name", category_slug)

        content += f"### {cat_name}\n\n"
        for api in sorted(categorized_apis[category_slug], key=lambda x: x["name"]):
            content += f"- [{api['name']}]({category_slug}/{api['id']}.md) - `{api['api_id']}`\n"
        content += "\n"

    # Write main index
    main_index_file = API_DOCS_DIR.parent / "api-index.md"
    with open(main_index_file, "w", encoding="utf-8") as f:
        f.write(content)

    logger.info(f"Generated main API index: {main_index_file}")


def main() -> None:
    """Main function."""
    parser = argparse.ArgumentParser(
        description="Parse KIS OpenAPI Excel and generate documentation"
    )
    parser.add_argument(
        "--generate-docs",
        action="store_true",
        help="Generate API documentation files"
    )
    parser.add_argument(
        "--phase",
        type=int,
        choices=[1, 2],
        help="Generation phase: 1 = priority APIs, 2 = remaining APIs"
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Parse Excel and generate documentation"
    )
    parser.add_argument(
        "--priority-categories",
        type=str,
        help="Comma-separated list of priority category slugs for phase 1"
    )

    args = parser.parse_args()

    try:
        # Default priority categories for phase 1
        default_priority = [
            "oauth-authentication",
            "domestic-stock-orders",
            "domestic-stock-realtime",
            "overseas-stock-orders",
            "overseas-stock-realtime",
        ]

        if args.priority_categories:
            priority_categories = args.priority_categories.split(",")
        else:
            priority_categories = default_priority

        # Always parse Excel to get latest data
        logger.info("=" * 60)
        logger.info("Step 1: Parsing Excel file")
        logger.info("=" * 60)
        tr_id_mapping = parse_excel()

        # Validate schema
        validate_json_schema(tr_id_mapping)

        # Save JSON
        save_json(tr_id_mapping)

        # Generate documentation if requested
        if args.generate_docs or args.all:
            logger.info("")
            logger.info("=" * 60)
            logger.info("Step 2: Generating API documentation")
            logger.info("=" * 60)
            generate_api_docs(tr_id_mapping, phase=args.phase, priority_categories=priority_categories)

            # Generate main index
            generate_main_index(tr_id_mapping)

        logger.info("")
        logger.info("=" * 60)
        logger.info("SUCCESS: All tasks completed!")
        logger.info("=" * 60)

    except Exception as e:
        logger.error(f"ERROR: {e}")
        raise


if __name__ == "__main__":
    main()
